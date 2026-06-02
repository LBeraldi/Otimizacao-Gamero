"""
sewer_optimizer.py  —  Otimizador de Redes Coletoras de Esgoto
==============================================================
Versão 2.0 — MVP comercial

Extensão da dissertação Gameiro (2003):
  "Dimensionamento otimizado de redes de esgotos usando Algoritmos Genéticos"

Melhorias sobre o SwOPy.py original:
  ✔ Topologia dinâmica via CSV (rede qualquer — remove hardcode)
  ✔ Cotas de terreno reais por nó (não mais todos em 200 m)
  ✔ Tabelas de custo e parâmetros via JSON (configurável por região/ano)
  ✔ Restrições hidráulicas completas:
        y/D ≤ 0.75  |  τ ≥ 1.0 Pa  |  v ≤ 5.0 m/s
        cobertura ≥ 1.20 m  |  profundidade ≤ 6.0 m
  ✔ Múltiplas execuções do AG com estatísticas (média, dp, melhor)
  ✔ Saída em Excel (.xlsx) com planilha hidráulica completa
  ✔ CLI com argparse:
        python sewer_optimizer.py --network rede.csv --config config.json

Uso rápido (gera arquivos de exemplo e roda):
    python sewer_optimizer.py --generate-samples
    python sewer_optimizer.py --network sample_network.csv --config sample_config.json
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import random
import sys
import statistics
from collections import defaultdict, deque

# Garante saída UTF-8 no Windows (evita UnicodeEncodeError no console)
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ─────────────────────────────────────────────────────────────────────────────
# DATACLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class Edge:
    """Um trecho da rede (coletor entre dois poços de visita)."""
    row_id: int
    label: str
    upstream: int           # nó de montante
    downstream: int         # nó de jusante
    length_m: float
    n_manning: float        # coeficiente de rugosidade de Manning
    q_ini_m3s: float        # vazão inicial (ponta de rede)
    q_fin_m3s: float        # vazão final (contribuição acumulada)
    ground_up_m: float      # cota do terreno no nó de montante  [m]
    ground_down_m: float    # cota do terreno no nó de jusante   [m]


@dataclass
class ProjectConfig:
    """Todos os parâmetros configuráveis do projeto."""
    # GA
    population_size: int = 300
    n_generations: int = 1000
    crossover_rate: float = 0.90
    mutation_rate: float = 0.30
    elite_count: int = 3
    n_runs: int = 5
    random_seed: int = 42

    # Hidráulica
    max_y_over_d: float = 0.75
    min_shear_pa: float = 1.0
    max_velocity_ms: float = 5.0
    min_cover_m: float = 1.20
    max_trench_depth_m: float = 6.0
    initial_cover_m: float = 1.50  # profundidade mínima na ponta de rede

    # Conjuntos discretos
    discrete_diameters_mm: List[int] = field(default_factory=lambda: [150,200,250,300,350,400,450])
    discrete_slopes: List[float] = field(default_factory=lambda: [0.003,0.004,0.005,0.006,0.007,0.008,0.009])
    depth_classes: List[float] = field(default_factory=lambda: [2.0,3.0,4.5,6.0,8.0])

    # Custos
    currency: str = "USD"
    cost_year: int = 1995
    a_ij: Dict = field(default_factory=dict)   # coeficientes polinomiais do coletor
    pv_cost_table: Dict = field(default_factory=dict)  # custos dos PVs

    # Penalidades
    k_penalty: float = 0.8
    p1_y_over_d: float = 500_000.0
    p2_shear: float = 50_000.0
    p3_velocity: float = 100_000.0
    p4_depth: float = 200_000.0
    p5_cover: float = 200_000.0


@dataclass
class EdgeResult:
    edge: Edge
    diameter_mm: int
    slope: float
    uc_up_m: float          # cota do invert de montante
    dc_down_m: float        # cota do invert de jusante
    depth_up_m: float       # profundidade do PV de montante
    depth_down_m: float     # profundidade do PV de jusante
    depth_class_m: float    # classe de profundidade para custo
    y_over_d_ini: float
    y_over_d_fin: float
    v_ini: float
    v_fin: float
    v_max: float            # velocidade crítica
    shear_pa: float
    collector_cost: float
    upstream_pv_cost: float
    row_total: float
    # Flags de violação
    viol_y_over_d: bool = False
    viol_shear: bool = False
    viol_velocity: bool = False
    viol_depth: bool = False
    viol_cover: bool = False


@dataclass
class EvaluationResult:
    total_cost: float
    final_pv_cost: float
    # Somas de violações (para penalidade)
    excess_y_over_d: float
    deficit_shear: float
    excess_velocity: float
    excess_depth: float
    deficit_cover: float
    rows: List[EdgeResult]

    @property
    def is_feasible(self) -> bool:
        return (
            self.excess_y_over_d < 1e-9
            and self.deficit_shear < 1e-9
            and self.excess_velocity < 1e-9
            and self.excess_depth < 1e-9
            and self.deficit_cover < 1e-9
        )


@dataclass
class GARunResult:
    run_id: int
    seed_used: int
    best_chromosome: List[int]
    best_eval: EvaluationResult
    best_penalized_cost: float
    best_penalty: float
    best_generation: int
    history_best: List[float]


@dataclass
class MultiRunStats:
    n_runs: int
    best_run: GARunResult               # execução com menor custo base viável
    all_runs: List[GARunResult]
    feasible_runs: List[GARunResult]    # apenas execuções sem violações
    mean_cost: Optional[float]          # mean de execuções viáveis
    std_cost: Optional[float]
    min_cost: Optional[float]
    max_cost: Optional[float]


# ─────────────────────────────────────────────────────────────────────────────
# CARREGAMENTO DE CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

_DEFAULT_A_IJ = {
    0: {0: -29.3185,  1: 28.92595,  2: -5.42286,   3:  0.524836},
    1: {0:  0.338185, 1:  0.045221, 2: -0.0063,     3:  7.81e-04},
    2: {0: -0.00149,  1: -1.15e-04, 2:  1.66e-05,   3: -1.97e-06},
    3: {0:  3.29e-06, 1:  1.60e-07, 2: -2.30e-08,   3:  2.75e-09},
}

_DEFAULT_PV = {
    100:  {2.0: 556.0,  3.0:  556.0, 4.5: 1461.0, 6.0: 1636.0, 8.0: 2054.0},
    150:  {2.0: 556.0,  3.0:  556.0, 4.5: 1461.0, 6.0: 1636.0, 8.0: 2054.0},
    200:  {2.0: 769.0,  3.0:  844.0, 4.5: 1461.0, 6.0: 1636.0, 8.0: 2054.0},
    250:  {2.0: 769.0,  3.0:  844.0, 4.5: 1461.0, 6.0: 1636.0, 8.0: 2054.0},
    300:  {2.0: 769.0,  3.0:  844.0, 4.5: 1461.0, 6.0: 1636.0, 8.0: 2054.0},
    350:  {2.0: 1131.0, 3.0: 1288.0, 4.5: 1633.0, 6.0: 1822.0, 8.0: 2282.0},
    400:  {2.0: 1131.0, 3.0: 1288.0, 4.5: 1633.0, 6.0: 1822.0, 8.0: 2282.0},
    450:  {2.0: 1131.0, 3.0: 1288.0, 4.5: 1633.0, 6.0: 1822.0, 8.0: 2282.0},
}


def load_config(config_path: Optional[Path]) -> ProjectConfig:
    """Carrega configuração de JSON; usa defaults se arquivo não fornecido."""
    cfg = ProjectConfig()
    cfg.a_ij = _DEFAULT_A_IJ
    cfg.pv_cost_table = _DEFAULT_PV

    if config_path is None:
        return cfg

    with config_path.open(encoding="utf-8") as f:
        data = json.load(f)

    ga = data.get("ga", {})
    cfg.population_size  = ga.get("population_size",  cfg.population_size)
    cfg.n_generations    = ga.get("n_generations",    cfg.n_generations)
    cfg.crossover_rate   = ga.get("crossover_rate",   cfg.crossover_rate)
    cfg.mutation_rate    = ga.get("mutation_rate",    cfg.mutation_rate)
    cfg.elite_count      = ga.get("elite_count",      cfg.elite_count)
    cfg.n_runs           = ga.get("n_runs",           cfg.n_runs)
    cfg.random_seed      = ga.get("random_seed",      cfg.random_seed)

    hyd = data.get("hydraulics", {})
    cfg.max_y_over_d       = hyd.get("max_y_over_d",       cfg.max_y_over_d)
    cfg.min_shear_pa       = hyd.get("min_shear_pa",        cfg.min_shear_pa)
    cfg.max_velocity_ms    = hyd.get("max_velocity_ms",     cfg.max_velocity_ms)
    cfg.min_cover_m        = hyd.get("min_cover_m",         cfg.min_cover_m)
    cfg.max_trench_depth_m = hyd.get("max_trench_depth_m",  cfg.max_trench_depth_m)
    cfg.initial_cover_m    = hyd.get("initial_cover_m",     cfg.initial_cover_m)

    if "discrete_diameters_mm" in data:
        cfg.discrete_diameters_mm = [int(d) for d in data["discrete_diameters_mm"]]
    if "discrete_slopes" in data:
        cfg.discrete_slopes = [float(s) for s in data["discrete_slopes"]]
    if "depth_classes" in data:
        cfg.depth_classes = [float(c) for c in data["depth_classes"]]

    costs = data.get("costs", {})
    cfg.currency  = costs.get("currency", cfg.currency)
    cfg.cost_year = costs.get("year",     cfg.cost_year)

    if "a_ij" in costs:
        # JSON tem chaves string; converte para int
        cfg.a_ij = {
            int(i): {int(j): float(v) for j, v in row.items()}
            for i, row in costs["a_ij"].items()
        }

    if "pv_cost_table" in costs:
        cfg.pv_cost_table = {
            int(diam): {float(cls): float(val) for cls, val in depths.items()}
            for diam, depths in costs["pv_cost_table"].items()
        }

    pen = data.get("penalties", {})
    cfg.k_penalty    = pen.get("k_exponent",   cfg.k_penalty)
    cfg.p1_y_over_d  = pen.get("p1_y_over_d",  cfg.p1_y_over_d)
    cfg.p2_shear     = pen.get("p2_shear",      cfg.p2_shear)
    cfg.p3_velocity  = pen.get("p3_velocity",   cfg.p3_velocity)
    cfg.p4_depth     = pen.get("p4_depth",      cfg.p4_depth)
    cfg.p5_cover     = pen.get("p5_cover",      cfg.p5_cover)

    return cfg


# ─────────────────────────────────────────────────────────────────────────────
# CARREGAMENTO DA REDE (CSV)
# ─────────────────────────────────────────────────────────────────────────────

def load_network_csv(network_path: Path) -> List[Edge]:
    """
    Lê a rede coletora de um CSV com as colunas:
        edge_id, label, upstream_node, downstream_node, length_m,
        n_manning, q_initial_m3s, q_final_m3s,
        ground_elev_upstream_m, ground_elev_downstream_m

    Colunas n_manning são opcionais — default 0.013 se ausentes.
    """
    edges: List[Edge] = []
    required = {
        "edge_id","label","upstream_node","downstream_node","length_m",
        "q_initial_m3s","q_final_m3s",
        "ground_elev_upstream_m","ground_elev_downstream_m",
    }

    with network_path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"CSV vazio ou sem cabeçalho: {network_path}")

        headers = {h.strip().lower() for h in reader.fieldnames}
        missing = required - headers
        if missing:
            raise ValueError(
                f"CSV de rede com colunas faltando: {missing}\n"
                f"Colunas encontradas: {headers}"
            )

        for lineno, row in enumerate(reader, start=2):
            try:
                n_manning = float(row.get("n_manning", 0.013) or 0.013)
                edges.append(Edge(
                    row_id       = int(row["edge_id"]),
                    label        = str(row["label"]).strip(),
                    upstream     = int(row["upstream_node"]),
                    downstream   = int(row["downstream_node"]),
                    length_m     = float(row["length_m"]),
                    n_manning    = n_manning,
                    q_ini_m3s   = float(row["q_initial_m3s"]),
                    q_fin_m3s   = float(row["q_final_m3s"]),
                    ground_up_m  = float(row["ground_elev_upstream_m"]),
                    ground_down_m= float(row["ground_elev_downstream_m"]),
                ))
            except (KeyError, ValueError) as exc:
                raise ValueError(f"Erro na linha {lineno} de {network_path}: {exc}") from exc

    if not edges:
        raise ValueError(f"Nenhum trecho encontrado em {network_path}")

    return edges


def validate_network(edges: List[Edge]) -> List[str]:
    """Retorna lista de avisos/erros de validação (lista vazia = OK)."""
    issues: List[str] = []
    ids = [e.row_id for e in edges]
    if len(ids) != len(set(ids)):
        issues.append("Existem edge_id duplicados na rede.")

    for e in edges:
        if e.length_m <= 0:
            issues.append(f"Trecho {e.row_id} ({e.label}): comprimento ≤ 0.")
        if e.q_fin_m3s < e.q_ini_m3s:
            issues.append(f"Trecho {e.row_id}: q_final < q_inicial (vazão decrescente).")
        if e.q_ini_m3s <= 0:
            issues.append(f"Trecho {e.row_id}: q_inicial ≤ 0.")
        if e.upstream == e.downstream:
            issues.append(f"Trecho {e.row_id}: upstream == downstream (laço).")

    # Verifica ciclos via DFS simples
    adj: Dict[int, List[int]] = defaultdict(list)
    for e in edges:
        adj[e.upstream].append(e.downstream)
    visited: set = set()
    stack: set = set()

    def has_cycle(node: int) -> bool:
        visited.add(node)
        stack.add(node)
        for nb in adj[node]:
            if nb not in visited:
                if has_cycle(nb):
                    return True
            elif nb in stack:
                return True
        stack.discard(node)
        return False

    all_nodes = {e.upstream for e in edges} | {e.downstream for e in edges}
    for n in all_nodes:
        if n not in visited:
            if has_cycle(n):
                issues.append("A rede contém ciclos — deve ser um DAG (grafo acíclico dirigido).")
                break

    return issues


# ─────────────────────────────────────────────────────────────────────────────
# TOPOLOGIA
# ─────────────────────────────────────────────────────────────────────────────

def build_topology(edges: List[Edge]) -> Tuple[Dict, Dict, List[int], int]:
    """
    Retorna:
        in_edges   — dict node -> [índices de trechos que chegam ao nó]
        out_edge   — dict node -> índice do trecho que sai do nó
        topo_order — lista de nós em ordem topológica (Kahn)
        outlet_node — nó terminal (sem saída)
    """
    in_edges: Dict[int, List[int]] = defaultdict(list)
    out_edge: Dict[int, int] = {}
    nodes: set = set()

    for idx, edge in enumerate(edges):
        in_edges[edge.downstream].append(idx)
        out_edge[edge.upstream] = idx
        nodes.add(edge.upstream)
        nodes.add(edge.downstream)

    indegree = {n: 0 for n in nodes}
    for edge in edges:
        indegree[edge.downstream] += 1

    queue = deque(sorted(n for n, d in indegree.items() if d == 0))
    topo_order: List[int] = []

    while queue:
        node = queue.popleft()
        topo_order.append(node)
        if node in out_edge:
            nxt = edges[out_edge[node]].downstream
            indegree[nxt] -= 1
            if indegree[nxt] == 0:
                queue.append(nxt)

    outlet_node = topo_order[-1]
    return dict(in_edges), out_edge, topo_order, outlet_node


# ─────────────────────────────────────────────────────────────────────────────
# FUNÇÕES HIDRÁULICAS
# ─────────────────────────────────────────────────────────────────────────────

def theta_from_y_over_d(y_over_d: float) -> float:
    y_over_d = max(1e-9, min(0.999999, y_over_d))
    return 2.0 * math.acos(1.0 - 2.0 * y_over_d)


def circular_section(diameter_m: float, theta: float) -> Tuple[float, float]:
    """Retorna (area_m2, raio_hidraulico_m) para seção circular parcialmente cheia."""
    area = (diameter_m ** 2 / 8.0) * (theta - math.sin(theta))
    perim = (diameter_m / 2.0) * theta
    rh = area / perim if perim > 0 else 0.0
    return area, rh


def flow_at_y_over_d(y_over_d: float, diam_m: float, slope: float, n: float) -> float:
    theta = theta_from_y_over_d(y_over_d)
    area, rh = circular_section(diam_m, theta)
    if area <= 0 or rh <= 0 or slope <= 0:
        return 0.0
    return (1.0 / n) * area * (rh ** (2.0 / 3.0)) * math.sqrt(slope)


def solve_y_over_d(flow_m3s: float, diam_m: float, slope: float, n: float) -> float:
    """Resolve y/D para uma dada vazão por bisecção (80 iterações)."""
    if flow_m3s <= 0.0:
        return 0.0
    lo, hi = 1e-8, 0.999999
    for _ in range(80):
        mid = (lo + hi) / 2.0
        if flow_at_y_over_d(mid, diam_m, slope, n) < flow_m3s:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2.0


# ─────────────────────────────────────────────────────────────────────────────
# CACHE HIDRÁULICO  (calculado após carregar a rede)
# ─────────────────────────────────────────────────────────────────────────────

HydKey = Tuple[int, int, float]   # (edge_idx, diameter_mm, slope)
HydVal = Tuple[float, float, float, float, float, float]
# (y_ini, y_fin, v_ini, v_fin, v_critical, shear_pa)

def build_hydraulic_cache(
    edges: List[Edge],
    diameters: List[int],
    slopes: List[float],
) -> Dict[HydKey, HydVal]:
    cache: Dict[HydKey, HydVal] = {}
    for idx, edge in enumerate(edges):
        for diam_mm in diameters:
            diam_m = diam_mm / 1000.0
            for slope in slopes:
                y_ini = solve_y_over_d(edge.q_ini_m3s, diam_m, slope, edge.n_manning)
                y_fin = solve_y_over_d(edge.q_fin_m3s, diam_m, slope, edge.n_manning)

                area_i, rh_i = circular_section(diam_m, theta_from_y_over_d(y_ini))
                area_f, rh_f = circular_section(diam_m, theta_from_y_over_d(y_fin))

                v_ini = edge.q_ini_m3s / area_i if area_i > 0 else 0.0
                v_fin = edge.q_fin_m3s / area_f if area_f > 0 else 0.0
                v_crit = 6.0 * math.sqrt(9.81 * rh_f) if rh_f > 0 else 0.0
                shear  = 9810.0 * rh_i * slope if rh_i > 0 else 0.0

                cache[(idx, diam_mm, slope)] = (y_ini, y_fin, v_ini, v_fin, v_crit, shear)
    return cache


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOS
# ─────────────────────────────────────────────────────────────────────────────

def ceil_depth_class(depth_m: float, classes: List[float]) -> float:
    for cls in sorted(classes):
        if depth_m <= cls + 1e-9:
            return cls
    return max(classes)


def collector_unit_cost(diam_mm: int, depth_cls: float, a_ij: Dict) -> float:
    """Custo unitário do coletor ($/m) via regressão polinomial bidimensional."""
    total = 0.0
    for i, row in a_ij.items():
        for j, coef in row.items():
            total += coef * (diam_mm ** i) * (depth_cls ** j)
    return total


def pv_cost(depth_m: float, diam_mm: int, pv_table: Dict) -> float:
    """Custo do poço de visita via tabela de lookup (diâmetro × classe de profundidade)."""
    # Encontra chave de diâmetro mais próxima por baixo
    avail_diams = sorted(pv_table.keys())
    key_diam = avail_diams[0]
    for d in avail_diams:
        if diam_mm >= d:
            key_diam = d

    depth_costs = pv_table[key_diam]
    for cls in sorted(depth_costs):
        if depth_m <= cls + 1e-9:
            return depth_costs[cls]
    return depth_costs[max(sorted(depth_costs))]


# ─────────────────────────────────────────────────────────────────────────────
# CROMOSSOMO
# cromossomo = [d0,s0, d1,s1, ..., dN,sN]  — códigos 0..len-1
# ─────────────────────────────────────────────────────────────────────────────

def random_chromosome(n_edges: int, n_diams: int, n_slopes: int) -> List[int]:
    return [
        random.randrange(n_diams) if i % 2 == 0 else random.randrange(n_slopes)
        for i in range(2 * n_edges)
    ]


def decode_chromosome(
    chromosome: List[int],
    diameters: List[int],
    slopes: List[float],
) -> List[Tuple[int, float]]:
    genes = []
    for i in range(0, len(chromosome), 2):
        genes.append((diameters[chromosome[i]], slopes[chromosome[i + 1]]))
    return genes


# ─────────────────────────────────────────────────────────────────────────────
# AVALIAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

def evaluate(
    chromosome: List[int],
    edges: List[Edge],
    cfg: ProjectConfig,
    in_edges: Dict,
    out_edge: Dict,
    topo_order: List[int],
    outlet_node: int,
    hyd_cache: Dict,
    eval_cache: Dict,
) -> EvaluationResult:
    key = tuple(chromosome)
    if key in eval_cache:
        return eval_cache[key]

    genes = decode_chromosome(chromosome, cfg.discrete_diameters_mm, cfg.discrete_slopes)
    raw: List[Optional[Dict]] = [None] * len(edges)

    for node in topo_order:
        if node not in out_edge:
            continue
        eidx = out_edge[node]
        edge = edges[eidx]
        diam_mm, slope = genes[eidx]

        # Cota do invert de montante
        if in_edges.get(edge.upstream):
            # Pega o mínimo invert de jusante de todos os trechos que chegam
            uc_up = min(raw[ii]["dc_down_m"] for ii in in_edges[edge.upstream])
        else:
            uc_up = edge.ground_up_m - cfg.initial_cover_m

        dc_down = uc_up - slope * edge.length_m
        depth_up   = edge.ground_up_m   - uc_up
        depth_down = edge.ground_down_m - dc_down
        mean_depth = (depth_up + depth_down) / 2.0
        depth_cls  = ceil_depth_class(mean_depth, cfg.depth_classes)

        y_ini, y_fin, v_ini, v_fin, v_crit, shear = hyd_cache[(eidx, diam_mm, slope)]
        col_cost = collector_unit_cost(diam_mm, depth_cls, cfg.a_ij) * edge.length_m

        raw[eidx] = {
            "edge": edge, "diam_mm": diam_mm, "slope": slope,
            "uc_up_m": uc_up, "dc_down_m": dc_down,
            "depth_up_m": depth_up, "depth_down_m": depth_down, "depth_cls_m": depth_cls,
            "y_ini": y_ini, "y_fin": y_fin,
            "v_ini": v_ini, "v_fin": v_fin, "v_crit": v_crit, "shear": shear,
            "col_cost": col_cost,
        }

    # Acumula profundidades e diâmetros máximos por nó (para custos dos PVs)
    node_max_depth: Dict[int, float]  = defaultdict(float)
    node_max_diam:  Dict[int, int]    = defaultdict(int)
    for r in raw:
        node_max_depth[r["edge"].upstream]   = max(node_max_depth[r["edge"].upstream],   r["depth_up_m"])
        node_max_depth[r["edge"].downstream] = max(node_max_depth[r["edge"].downstream], r["depth_down_m"])
        node_max_diam[r["edge"].upstream]    = max(node_max_diam[r["edge"].upstream],    r["diam_mm"])
        node_max_diam[r["edge"].downstream]  = max(node_max_diam[r["edge"].downstream],  r["diam_mm"])

    rows: List[EdgeResult] = []
    pv_sum = 0.0

    # Acumuladores de violação
    exc_y = exc_v = exc_d = 0.0
    def_s = def_c = 0.0

    for r in raw:
        e = r["edge"]
        up_pv = pv_cost(node_max_depth[e.upstream], node_max_diam[e.upstream], cfg.pv_cost_table)
        pv_sum += up_pv

        v_worst = max(r["v_ini"], r["v_fin"])
        d_worst = max(r["depth_up_m"], r["depth_down_m"])
        c_worst = min(r["depth_up_m"], r["depth_down_m"])  # cobertura mínima = menor profundidade

        viol_y  = r["y_fin"]  > cfg.max_y_over_d   + 1e-9
        viol_s  = r["shear"]  < cfg.min_shear_pa   - 1e-9
        viol_v  = v_worst     > cfg.max_velocity_ms + 1e-9
        viol_d  = d_worst     > cfg.max_trench_depth_m + 1e-9
        viol_c  = c_worst     < cfg.min_cover_m    - 1e-9

        exc_y += max(r["y_fin"]  - cfg.max_y_over_d,    0.0)
        def_s += max(cfg.min_shear_pa   - r["shear"],   0.0)
        exc_v += max(v_worst - cfg.max_velocity_ms,      0.0)
        exc_d += max(d_worst - cfg.max_trench_depth_m,  0.0)
        def_c += max(cfg.min_cover_m - c_worst,          0.0)

        rows.append(EdgeResult(
            edge=e, diameter_mm=r["diam_mm"], slope=r["slope"],
            uc_up_m=r["uc_up_m"], dc_down_m=r["dc_down_m"],
            depth_up_m=r["depth_up_m"], depth_down_m=r["depth_down_m"],
            depth_class_m=r["depth_cls_m"],
            y_over_d_ini=r["y_ini"], y_over_d_fin=r["y_fin"],
            v_ini=r["v_ini"], v_fin=r["v_fin"], v_max=r["v_crit"],
            shear_pa=r["shear"],
            collector_cost=r["col_cost"],
            upstream_pv_cost=up_pv,
            row_total=r["col_cost"] + up_pv,
            viol_y_over_d=viol_y, viol_shear=viol_s,
            viol_velocity=viol_v, viol_depth=viol_d, viol_cover=viol_c,
        ))

    final_pv = pv_cost(node_max_depth[outlet_node], node_max_diam[outlet_node], cfg.pv_cost_table)
    total = sum(r.collector_cost for r in rows) + pv_sum + final_pv

    result = EvaluationResult(
        total_cost=total, final_pv_cost=final_pv,
        excess_y_over_d=exc_y, deficit_shear=def_s,
        excess_velocity=exc_v, excess_depth=exc_d, deficit_cover=def_c,
        rows=rows,
    )
    eval_cache[key] = result
    return result


def penalized_cost(
    chromosome: List[int], generation: int, cfg: ProjectConfig,
    edges, in_edges, out_edge, topo_order, outlet_node, hyd_cache, eval_cache,
) -> Tuple[float, EvaluationResult, float]:
    base = evaluate(chromosome, edges, cfg, in_edges, out_edge, topo_order, outlet_node, hyd_cache, eval_cache)
    gf = (generation / cfg.n_generations) ** cfg.k_penalty
    penalty = gf * (
        cfg.p1_y_over_d * base.excess_y_over_d
        + cfg.p2_shear   * base.deficit_shear
        + cfg.p3_velocity * base.excess_velocity
        + cfg.p4_depth    * base.excess_depth
        + cfg.p5_cover    * base.deficit_cover
    )
    return base.total_cost + penalty, base, penalty


# ─────────────────────────────────────────────────────────────────────────────
# ALGORITMO GENÉTICO
# ─────────────────────────────────────────────────────────────────────────────

def aptitudes(costs: List[float]) -> List[float]:
    worst = max(costs)
    return [max(worst * 1.2 - c, 1e-9) for c in costs]


def roulette_select(population: List[List[int]], apts: List[float]) -> List[int]:
    total = sum(apts)
    r = random.random() * total
    acc = 0.0
    for chrom, apt in zip(population, apts):
        acc += apt
        if acc >= r:
            return chrom.copy()
    return population[-1].copy()


def crossover(pa: List[int], pb: List[int], rate: float) -> Tuple[List[int], List[int]]:
    if random.random() >= rate:
        return pa.copy(), pb.copy()
    cut = random.randint(1, len(pa) - 1)
    return pa[:cut] + pb[cut:], pb[:cut] + pa[cut:]


def mutate(chrom: List[int], rate: float, n_diams: int, n_slopes: int) -> None:
    if random.random() >= rate:
        return
    idx = random.randrange(len(chrom))
    n_opts = n_diams if idx % 2 == 0 else n_slopes
    cur = chrom[idx]
    opts = [x for x in range(n_opts) if x != cur]
    chrom[idx] = random.choice(opts)


def run_ga(
    edges: List[Edge],
    cfg: ProjectConfig,
    in_edges: Dict, out_edge: Dict,
    topo_order: List[int], outlet_node: int,
    hyd_cache: Dict,
    run_id: int = 1,
    seed: Optional[int] = None,
    verbose: bool = True,
) -> GARunResult:
    seed_used = seed if seed is not None else cfg.random_seed
    random.seed(seed_used)

    n_e = len(edges)
    n_d = len(cfg.discrete_diameters_mm)
    n_s = len(cfg.discrete_slopes)
    eval_cache: Dict = {}

    population = [random_chromosome(n_e, n_d, n_s) for _ in range(cfg.population_size)]
    history: List[float] = []

    best_chrom: List[int] = []
    best_eval_r: Optional[EvaluationResult] = None
    best_pen = float("inf")
    best_penalty = 0.0
    best_gen = 1

    for gen in range(1, cfg.n_generations + 1):
        evaluated = [
            penalized_cost(ch, gen, cfg, edges, in_edges, out_edge, topo_order, outlet_node, hyd_cache, eval_cache)
            for ch in population
        ]
        ranked = sorted(zip(population, evaluated), key=lambda x: x[1][0])
        population    = [ch for ch, _ in ranked]
        pen_vals      = [t[0] for _, t in ranked]
        base_evals    = [t[1] for _, t in ranked]
        penalties     = [t[2] for _, t in ranked]

        if pen_vals[0] < best_pen:
            best_pen     = pen_vals[0]
            best_chrom   = population[0].copy()
            best_eval_r  = base_evals[0]
            best_penalty = penalties[0]
            best_gen     = gen

        history.append(pen_vals[0])

        if verbose and (gen == 1 or gen % 100 == 0):
            feas = "OK" if base_evals[0].is_feasible else "!!"
            print(
                f"  [run {run_id}] gen {gen:5d} | pen={pen_vals[0]:>12,.2f} "
                f"| custo={base_evals[0].total_cost:>12,.2f} "
                f"| penalidade={penalties[0]:>10,.2f} [{feas}]"
            )

        apts_ = aptitudes(pen_vals)
        new_pop = [population[i].copy() for i in range(cfg.elite_count)]

        while len(new_pop) < cfg.population_size:
            pa = roulette_select(population, apts_)
            pb = roulette_select(population, apts_)
            ca, cb = crossover(pa, pb, cfg.crossover_rate)
            mutate(ca, cfg.mutation_rate, n_d, n_s)
            mutate(cb, cfg.mutation_rate, n_d, n_s)
            new_pop.append(ca)
            if len(new_pop) < cfg.population_size:
                new_pop.append(cb)

        population = new_pop

    return GARunResult(
        run_id=run_id, seed_used=seed_used,
        best_chromosome=best_chrom, best_eval=best_eval_r,
        best_penalized_cost=best_pen, best_penalty=best_penalty,
        best_generation=best_gen, history_best=history,
    )


# ─────────────────────────────────────────────────────────────────────────────
# MÚLTIPLAS EXECUÇÕES
# ─────────────────────────────────────────────────────────────────────────────

def run_multiple(
    edges, cfg, in_edges, out_edge, topo_order, outlet_node, hyd_cache,
    verbose: bool = True,
) -> MultiRunStats:
    all_runs: List[GARunResult] = []
    for i in range(cfg.n_runs):
        seed = cfg.random_seed + i
        print(f"\n{'='*60}")
        print(f"EXECUÇÃO {i+1}/{cfg.n_runs}  (seed={seed})")
        print(f"{'='*60}")
        r = run_ga(edges, cfg, in_edges, out_edge, topo_order, outlet_node, hyd_cache,
                   run_id=i+1, seed=seed, verbose=verbose)
        all_runs.append(r)
        feas_str = "VIAVEL" if r.best_eval.is_feasible else "INVIAVEL"
        print(f"  >> Melhor custo base: {r.best_eval.total_cost:,.2f} {cfg.currency}  ({feas_str})")

    feasible = [r for r in all_runs if r.best_eval.is_feasible]
    # Se não há viáveis, usa todas para estatísticas
    pool = feasible if feasible else all_runs
    costs = [r.best_eval.total_cost for r in pool]

    mean_c = statistics.mean(costs)    if len(costs) > 1 else costs[0]
    std_c  = statistics.stdev(costs)   if len(costs) > 1 else 0.0
    min_c  = min(costs)
    max_c  = max(costs)

    best_run = min(pool, key=lambda r: r.best_eval.total_cost)

    return MultiRunStats(
        n_runs=cfg.n_runs, best_run=best_run, all_runs=all_runs,
        feasible_runs=feasible,
        mean_cost=mean_c, std_cost=std_c, min_cost=min_c, max_cost=max_c,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RELATÓRIOS — CONSOLE
# ─────────────────────────────────────────────────────────────────────────────

def print_run_summary(stats: MultiRunStats, cfg: ProjectConfig) -> None:
    cur = cfg.currency
    print(f"\n{'='*70}")
    print("RESUMO DAS MÚLTIPLAS EXECUÇÕES")
    print(f"{'='*70}")
    print(f"  Execuções totais    : {stats.n_runs}")
    print(f"  Execuções viáveis   : {len(stats.feasible_runs)}")
    if stats.feasible_runs:
        print(f"  Custo mínimo (viável)  : {stats.min_cost:>14,.2f} {cur}")
        print(f"  Custo máximo (viável)  : {stats.max_cost:>14,.2f} {cur}")
        print(f"  Custo médio  (viável)  : {stats.mean_cost:>14,.2f} {cur}")
        print(f"  Desvio padrão          : {stats.std_cost:>14,.2f} {cur}")
    else:
        print("  ATENÇÃO: nenhuma execução gerou solução hidraulicamente viável.")
        print(f"  Custo mínimo (inviável): {stats.min_cost:>14,.2f} {cur}")

    print(f"\nMELHOR SOLUÇÃO (execução {stats.best_run.run_id}, seed {stats.best_run.seed_used})")
    print(f"  Custo total         : {stats.best_run.best_eval.total_cost:>14,.2f} {cur}")
    print(f"  Custo penalizado    : {stats.best_run.best_penalized_cost:>14,.2f} {cur}")
    print(f"  Penalidade          : {stats.best_run.best_penalty:>14,.2f} {cur}")
    print(f"  Geração do melhor   : {stats.best_run.best_generation}")
    ev = stats.best_run.best_eval
    print(f"  Viavel              : {'SIM [OK]' if ev.is_feasible else 'NAO [!!]'}")
    if not ev.is_feasible:
        print(f"    excesso y/D       : {ev.excess_y_over_d:.6f}")
        print(f"    déficit shear     : {ev.deficit_shear:.6f}")
        print(f"    excesso velocidade: {ev.excess_velocity:.6f}")
        print(f"    excesso profund.  : {ev.excess_depth:.6f}")
        print(f"    déficit cobertura : {ev.deficit_cover:.6f}")

    print(f"\n{'─'*70}")
    print("PLANILHA HIDRÁULICA DA MELHOR SOLUÇÃO")
    print(f"{'─'*70}")
    hdr = (
        f"{'Id':>3} {'Rótulo':<7} {'M->J':>7} {'L':>5} {'D':>4} {'S':>7} "
        f"{'hM':>6} {'hJ':>6} {'hCls':>5} "
        f"{'y/Di':>6} {'y/Df':>6} {'tau':>8} {'vMax':>6} "
        f"{'C_col':>11} {'C_PV':>8} {'Total':>11}"
    )
    print(hdr)
    print("─" * len(hdr))
    for r in ev.rows:
        flags = ""
        if r.viol_y_over_d: flags += "Y"
        if r.viol_shear:     flags += "τ"
        if r.viol_velocity:  flags += "V"
        if r.viol_depth:     flags += "D"
        if r.viol_cover:     flags += "C"
        flags_str = f" [{flags}]" if flags else ""
        print(
            f"{r.edge.row_id:>3} "
            f"{r.edge.label:<7} "
            f"{r.edge.upstream:>3}->{r.edge.downstream:<3} "
            f"{r.edge.length_m:>5.0f} "
            f"{r.diameter_mm:>4d} "
            f"{r.slope:>7.4f} "
            f"{r.depth_up_m:>6.2f} "
            f"{r.depth_down_m:>6.2f} "
            f"{r.depth_class_m:>5.1f} "
            f"{r.y_over_d_ini:>6.3f} "
            f"{r.y_over_d_fin:>6.3f} "
            f"{r.shear_pa:>8.4f} "
            f"{max(r.v_ini, r.v_fin):>6.3f} "
            f"{r.collector_cost:>11,.2f} "
            f"{r.upstream_pv_cost:>8,.2f} "
            f"{r.row_total:>11,.2f}"
            f"{flags_str}"
        )
    print(f"\n  PV final (nó saída): {ev.final_pv_cost:>11,.2f} {cur}")
    print(f"  TOTAL              : {ev.total_cost:>11,.2f} {cur}")
    if any(r.viol_y_over_d or r.viol_shear or r.viol_velocity or r.viol_depth or r.viol_cover for r in ev.rows):
        print("\n  Legenda de violações: Y=lâmina  τ=tensão  V=velocidade  D=profundidade  C=cobertura")


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAÇÃO EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(stats: MultiRunStats, cfg: ProjectConfig, output_dir: Path) -> Optional[Path]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [aviso] openpyxl não instalado — exportação Excel ignorada.")
        print("          Instale com: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()
    cur = cfg.currency

    # ── Estilos ──────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    sub_fill  = PatternFill("solid", fgColor="2E75B6")
    sub_font  = Font(bold=True, color="FFFFFF", size=10)
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")   # verde claro
    warn_fill = PatternFill("solid", fgColor="FFEB9C")   # amarelo
    err_fill  = PatternFill("solid", fgColor="FFC7CE")   # vermelho claro
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header_row(ws, row_num, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def style_subheader_row(ws, row_num, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = sub_fill
            cell.font = sub_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # ── Aba 1: Resumo geral ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["OTIMIZADOR DE REDES COLETORAS — SEWER OPTIMIZER v2.0"])
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append([])
    ws1.append(["Parâmetros do Projeto"])
    ws1["A3"].font = Font(bold=True, size=11)
    for k, v in [
        ("Nº de trechos", len(stats.best_run.best_eval.rows)),
        ("Diâmetros discretos (mm)", str(cfg.discrete_diameters_mm)),
        ("Declividades discretas", str(cfg.discrete_slopes)),
        ("Moeda", cur),
        ("Ano de referência de custos", cfg.cost_year),
        ("Cobertura mínima (m)", cfg.min_cover_m),
        ("Profundidade máxima (m)", cfg.max_trench_depth_m),
        ("Velocidade máxima (m/s)", cfg.max_velocity_ms),
        ("y/D máximo", cfg.max_y_over_d),
        ("τ mínimo (Pa)", cfg.min_shear_pa),
    ]:
        ws1.append([k, v])

    ws1.append([])
    ws1.append(["Resultados das Execuções do AG"])
    ws1[f"A{ws1.max_row}"].font = Font(bold=True, size=11)
    ws1.append([
        "Execução", "Seed", "Melhor Custo Base", "Penalidade",
        "Custo Penalizado", "Geração do Melhor", "Viável?"
    ])
    style_subheader_row(ws1, ws1.max_row, 7)
    for r in stats.all_runs:
        ws1.append([
            r.run_id, r.seed_used,
            round(r.best_eval.total_cost, 2),
            round(r.best_penalty, 2),
            round(r.best_penalized_cost, 2),
            r.best_generation,
            "SIM" if r.best_eval.is_feasible else "NÃO",
        ])
        row_n = ws1.max_row
        fill = ok_fill if r.best_eval.is_feasible else err_fill
        for c in range(1, 8):
            ws1.cell(row=row_n, column=c).fill = fill
            ws1.cell(row=row_n, column=c).border = thin_border

    ws1.append([])
    ws1.append(["Estatísticas (execuções viáveis)" if stats.feasible_runs else "Estatísticas (todas — sem viáveis)"])
    ws1[f"A{ws1.max_row}"].font = Font(bold=True)
    for k, v in [
        (f"Custo mínimo ({cur})", round(stats.min_cost, 2)),
        (f"Custo máximo ({cur})", round(stats.max_cost, 2)),
        (f"Custo médio ({cur})",  round(stats.mean_cost, 2)),
        (f"Desvio padrão ({cur})", round(stats.std_cost, 2)),
    ]:
        ws1.append([k, v])

    auto_width(ws1)

    # ── Aba 2: Planilha hidráulica da melhor solução ──────────────────────────
    ws2 = wb.create_sheet("Planilha Hidráulica")
    ev = stats.best_run.best_eval

    ws2.append([f"Melhor Solução — Execução {stats.best_run.run_id} — Seed {stats.best_run.seed_used}"])
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.append([f"Custo total: {ev.total_cost:,.2f} {cur}  |  Viável: {'SIM' if ev.is_feasible else 'NÃO'}"])
    ws2.append([])

    hdrs = [
        "Id","Rótulo","Montante","Jusante","L (m)",
        "D (mm)","S (m/m)",
        "Cota Inv. M (m)","Cota Inv. J (m)",
        "Prof. M (m)","Prof. J (m)","Classe Prof. (m)",
        "y/D ini","y/D fin","v ini (m/s)","v fin (m/s)","v crít (m/s)",
        "τ (Pa)","C_col","C_PV_mont","Total trecho",
        "Viola y/D","Viola τ","Viola v","Viola Prof.","Viola Cob."
    ]
    ws2.append(hdrs)
    style_header_row(ws2, ws2.max_row, len(hdrs))

    for r in ev.rows:
        row_data = [
            r.edge.row_id, r.edge.label, r.edge.upstream, r.edge.downstream, r.edge.length_m,
            r.diameter_mm, r.slope,
            round(r.uc_up_m, 3), round(r.dc_down_m, 3),
            round(r.depth_up_m, 3), round(r.depth_down_m, 3), r.depth_class_m,
            round(r.y_over_d_ini, 4), round(r.y_over_d_fin, 4),
            round(r.v_ini, 4), round(r.v_fin, 4), round(r.v_max, 4),
            round(r.shear_pa, 5),
            round(r.collector_cost, 2), round(r.upstream_pv_cost, 2), round(r.row_total, 2),
            "SIM" if r.viol_y_over_d else "",
            "SIM" if r.viol_shear    else "",
            "SIM" if r.viol_velocity else "",
            "SIM" if r.viol_depth    else "",
            "SIM" if r.viol_cover    else "",
        ]
        ws2.append(row_data)
        rn = ws2.max_row
        has_viol = r.viol_y_over_d or r.viol_shear or r.viol_velocity or r.viol_depth or r.viol_cover
        row_fill = err_fill if has_viol else ok_fill
        for c in range(1, len(hdrs) + 1):
            cell = ws2.cell(row=rn, column=c)
            cell.border = thin_border
            if has_viol:
                # Só colorir colunas de violação
                if c >= 22:
                    cell.fill = err_fill
                else:
                    cell.fill = warn_fill
            else:
                cell.fill = ok_fill

    ws2.append([])
    ws2.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                round(sum(r.collector_cost for r in ev.rows), 2),
                round(sum(r.upstream_pv_cost for r in ev.rows), 2),
                round(ev.total_cost, 2)])
    rn = ws2.max_row
    ws2.cell(row=rn, column=19).font = Font(bold=True)
    ws2.cell(row=rn, column=20).font = Font(bold=True)
    ws2.cell(row=rn, column=21).font = Font(bold=True)

    auto_width(ws2)

    # ── Aba 3: Convergência ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Convergência")
    ws3.append(["Geração"] + [f"Execução {r.run_id}" for r in stats.all_runs])
    style_header_row(ws3, 1, len(stats.all_runs) + 1)

    max_gen = max(len(r.history_best) for r in stats.all_runs)
    for g in range(max_gen):
        row = [g + 1]
        for r in stats.all_runs:
            row.append(round(r.history_best[g], 2) if g < len(r.history_best) else "")
        ws3.append(row)

    auto_width(ws3)

    # ── Aba 4: Parâmetros config ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Configuração AG")
    ws4.append(["Parâmetro", "Valor"])
    style_header_row(ws4, 1, 2)
    for k, v in [
        ("Tamanho da população", cfg.population_size),
        ("Número de gerações", cfg.n_generations),
        ("Taxa de cruzamento (Pc)", cfg.crossover_rate),
        ("Taxa de mutação (Pm)", cfg.mutation_rate),
        ("Elitismo (n)", cfg.elite_count),
        ("Número de execuções", cfg.n_runs),
        ("Seed base", cfg.random_seed),
        ("Expoente penalidade (k)", cfg.k_penalty),
        ("P1 – y/D", cfg.p1_y_over_d),
        ("P2 – tensão trativa", cfg.p2_shear),
        ("P3 – velocidade", cfg.p3_velocity),
        ("P4 – profundidade", cfg.p4_depth),
        ("P5 – cobertura", cfg.p5_cover),
    ]:
        ws4.append([k, v])
    auto_width(ws4)

    out_path = output_dir / "sewer_optimizer_resultado.xlsx"
    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAÇÃO CSV (fallback / complementar)
# ─────────────────────────────────────────────────────────────────────────────

def export_csv(stats: MultiRunStats, cfg: ProjectConfig, output_dir: Path) -> Path:
    out_path = output_dir / "sewer_optimizer_resultado.csv"
    ev = stats.best_run.best_eval

    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([
            "trecho_id","rotulo","montante","jusante","comprimento_m",
            "diametro_mm","declividade",
            "cota_inv_montante_m","cota_inv_jusante_m",
            "profundidade_montante_m","profundidade_jusante_m","classe_profundidade_m",
            "y_D_inicial","y_D_final","v_inicial_ms","v_final_ms","v_critica_ms",
            "tensao_trativa_Pa",
            "custo_coletor","custo_pv_montante","total_trecho",
            "viola_lamina","viola_arraste","viola_velocidade","viola_profundidade","viola_cobertura"
        ])
        for r in ev.rows:
            w.writerow([
                r.edge.row_id, r.edge.label, r.edge.upstream, r.edge.downstream, r.edge.length_m,
                r.diameter_mm, r.slope,
                round(r.uc_up_m, 3), round(r.dc_down_m, 3),
                round(r.depth_up_m, 3), round(r.depth_down_m, 3), r.depth_class_m,
                round(r.y_over_d_ini, 4), round(r.y_over_d_fin, 4),
                round(r.v_ini, 4), round(r.v_fin, 4), round(r.v_max, 4),
                round(r.shear_pa, 5),
                round(r.collector_cost, 2), round(r.upstream_pv_cost, 2), round(r.row_total, 2),
                "sim" if r.viol_y_over_d else "nao",
                "sim" if r.viol_shear    else "nao",
                "sim" if r.viol_velocity else "nao",
                "sim" if r.viol_depth    else "nao",
                "sim" if r.viol_cover    else "nao",
            ])

        w.writerow([])
        w.writerow(["resumo_chave", "resumo_valor"])
        w.writerow(["execucao_melhor", stats.best_run.run_id])
        w.writerow(["seed_melhor", stats.best_run.seed_used])
        w.writerow([f"custo_total_{cfg.currency}", round(ev.total_cost, 2)])
        w.writerow([f"custo_pv_final_{cfg.currency}", round(ev.final_pv_cost, 2)])
        w.writerow(["viavel", "sim" if ev.is_feasible else "nao"])
        w.writerow(["execucoes_viaveis", len(stats.feasible_runs)])
        if stats.mean_cost is not None:
            w.writerow([f"custo_medio_viaveis_{cfg.currency}", round(stats.mean_cost, 2)])
            w.writerow([f"desvio_padrao_{cfg.currency}", round(stats.std_cost, 2)])
            w.writerow([f"custo_minimo_{cfg.currency}", round(stats.min_cost, 2)])
            w.writerow([f"custo_maximo_{cfg.currency}", round(stats.max_cost, 2)])

    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# GERADORES DE ARQUIVOS DE EXEMPLO
# ─────────────────────────────────────────────────────────────────────────────

def generate_sample_network(path: Path) -> None:
    """
    Gera CSV de exemplo com a rede hipotética de Gameiro (18 trechos).

    Terreno praticamente plano (200 m), com leve caimento em direção ao exutório
    (nó 19). Isso reproduz as condições do estudo original e permite que o AG
    encontre soluções viáveis. Para projetos reais, substitua as cotas de terreno
    pelos valores obtidos em levantamento topográfico.

    Colunas: edge_id, label, upstream_node, downstream_node, length_m,
             n_manning, q_initial_m3s, q_final_m3s,
             ground_elev_upstream_m, ground_elev_downstream_m
    """
    # Cotas levemente decrescentes em direção ao exutório (nó 19 = cota mais baixa)
    # Variação máxima de 1.5 m na rede inteira — compatível com os slopes discretos
    rows = [
        #  id   lbl     up  dn   L      n      q_ini    q_fin   z_up   z_dn
        (1,"1-2",   1,  2, 100, 0.013, 0.0022, 0.0040, 200.50, 200.20),
        (2,"2-2",   3,  2, 100, 0.013, 0.0022, 0.0040, 200.60, 200.20),
        (3,"10-2",  2,  4,  80, 0.013, 0.0056, 0.0112, 200.20, 200.00),
        (4,"10-4",  4,  5,  90, 0.013, 0.0074, 0.0140, 200.00, 199.75),
        (5,"10-6",  5,  6,  80, 0.013, 0.0090, 0.0180, 199.75, 199.55),
        (6,"4-2",   7,  8,  90, 0.013, 0.0022, 0.0036, 200.40, 200.10),
        (7,"5-2",   9,  8,  90, 0.013, 0.0022, 0.0036, 200.45, 200.10),
        (8,"8-2",   8, 10,  80, 0.013, 0.0052, 0.0112, 200.10, 199.90),
        (9,"8-4",  10, 12,  60, 0.013, 0.0064, 0.0128, 199.90, 199.75),
        (10,"3-2", 11, 12,  80, 0.013, 0.0022, 0.0032, 200.30, 199.75),
        (11,"9-2", 12, 13,  80, 0.013, 0.0096, 0.0192, 199.75, 199.55),
        (12,"9-4", 13,  6,  80, 0.013, 0.0104, 0.0208, 199.55, 199.55),
        (13,"11-2", 6, 14,  80, 0.013, 0.0218, 0.0436, 199.55, 199.20),
        (14,"6-2", 15, 18,  90, 0.013, 0.0022, 0.0036, 200.60, 200.20),
        (15,"7-2", 16, 18,  90, 0.013, 0.0022, 0.0036, 200.55, 200.20),
        (16,"12-2",18, 17,  80, 0.013, 0.0052, 0.0104, 200.20, 200.00),
        (17,"12-4",17, 14,  70, 0.013, 0.0066, 0.0132, 200.00, 199.20),
        (18,"13-2",14, 19,  60, 0.013, 0.0296, 0.0592, 199.20, 199.00),
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "edge_id","label","upstream_node","downstream_node","length_m",
            "n_manning","q_initial_m3s","q_final_m3s",
            "ground_elev_upstream_m","ground_elev_downstream_m"
        ])
        for r in rows:
            w.writerow(r)
    print(f"  Arquivo de rede gerado: {path}")


def generate_sample_config(path: Path) -> None:
    """Gera JSON de configuração completo com todos os parâmetros."""
    config = {
        "_comentario": "Configuração do Sewer Optimizer — edite conforme seu projeto",
        "ga": {
            "population_size": 300,
            "n_generations": 1000,
            "crossover_rate": 0.90,
            "mutation_rate": 0.30,
            "elite_count": 3,
            "n_runs": 3,
            "random_seed": 42
        },
        "hydraulics": {
            "max_y_over_d": 0.75,
            "min_shear_pa": 1.0,
            "max_velocity_ms": 5.0,
            "min_cover_m": 1.20,
            "max_trench_depth_m": 6.0,
            "initial_cover_m": 1.50
        },
        "discrete_diameters_mm": [150, 200, 250, 300, 350, 400, 450],
        "discrete_slopes": [0.003, 0.004, 0.005, 0.006, 0.007, 0.008, 0.009],
        "depth_classes": [2.0, 3.0, 4.5, 6.0, 8.0],
        "costs": {
            "currency": "USD",
            "year": 1995,
            "_nota_a_ij": "Coeficientes da regressão polinomial: C_unit = sum(a_ij * D^i * h^j) em $/m",
            "a_ij": {
                "0": {"0": -29.3185,  "1": 28.92595,  "2": -5.42286,   "3": 0.524836},
                "1": {"0": 0.338185,  "1": 0.045221,  "2": -0.0063,    "3": 0.000781},
                "2": {"0": -0.00149,  "1": -0.000115, "2": 0.0000166,  "3": -0.00000197},
                "3": {"0": 0.00000329,"1": 0.00000016,"2": -0.000000023,"3": 0.00000000275}
            },
            "_nota_pv": "Custo dos poços de visita por diâmetro máximo e classe de profundidade",
            "pv_cost_table": {
                "100":  {"2.0": 556,  "3.0": 556,  "4.5": 1461, "6.0": 1636, "8.0": 2054},
                "150":  {"2.0": 556,  "3.0": 556,  "4.5": 1461, "6.0": 1636, "8.0": 2054},
                "200":  {"2.0": 769,  "3.0": 844,  "4.5": 1461, "6.0": 1636, "8.0": 2054},
                "250":  {"2.0": 769,  "3.0": 844,  "4.5": 1461, "6.0": 1636, "8.0": 2054},
                "300":  {"2.0": 769,  "3.0": 844,  "4.5": 1461, "6.0": 1636, "8.0": 2054},
                "350":  {"2.0": 1131, "3.0": 1288, "4.5": 1633, "6.0": 1822, "8.0": 2282},
                "400":  {"2.0": 1131, "3.0": 1288, "4.5": 1633, "6.0": 1822, "8.0": 2282},
                "450":  {"2.0": 1131, "3.0": 1288, "4.5": 1633, "6.0": 1822, "8.0": 2282}
            }
        },
        "penalties": {
            "_nota": "Pesos das penalidades — aumente para forçar restrições mais rígidas",
            "k_exponent": 0.8,
            "p1_y_over_d": 500000.0,
            "p2_shear": 50000.0,
            "p3_velocity": 100000.0,
            "p4_depth": 200000.0,
            "p5_cover": 200000.0
        }
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    print(f"  Arquivo de configuração gerado: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN / CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sewer Optimizer v2.0 — Otimizador de redes coletoras de esgoto por AG",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  # Gerar arquivos de exemplo e sair:
  python sewer_optimizer.py --generate-samples

  # Rodar com rede e configuração customizadas:
  python sewer_optimizer.py --network minha_rede.csv --config meu_config.json

  # Rodar com rede customizada e configuração padrão (Gameiro 2003):
  python sewer_optimizer.py --network minha_rede.csv

  # Rodar os arquivos de exemplo gerados:
  python sewer_optimizer.py --network sample_network.csv --config sample_config.json

Formato do CSV de rede (colunas obrigatórias):
  edge_id, label, upstream_node, downstream_node, length_m,
  q_initial_m3s, q_final_m3s, ground_elev_upstream_m, ground_elev_downstream_m
  (n_manning é opcional — default 0.013)
        """,
    )
    parser.add_argument("--network",          type=Path, help="CSV com os trechos da rede")
    parser.add_argument("--config",           type=Path, help="JSON de configuração (opcional)")
    parser.add_argument("--output",           type=Path, default=Path("resultados"), help="Diretório de saída (default: resultados)")
    parser.add_argument("--generate-samples", action="store_true", help="Gera CSVs/JSON de exemplo e sai")
    parser.add_argument("--no-excel",         action="store_true", help="Pula geração do Excel")
    parser.add_argument("--quiet",            action="store_true", help="Modo silencioso (sem progresso por geração)")
    args = parser.parse_args()

    # ── Gerar amostras ────────────────────────────────────────────────────────
    if args.generate_samples:
        base = Path(".")
        generate_sample_network(base / "sample_network.csv")
        generate_sample_config(base  / "sample_config.json")
        print("\nArquivos gerados. Rode agora:")
        print("  python sewer_optimizer.py --network sample_network.csv --config sample_config.json")
        return

    if args.network is None:
        parser.print_help()
        print("\nERRO: --network é obrigatório. Use --generate-samples para criar um exemplo.")
        sys.exit(1)

    # ── Carregar configuração ─────────────────────────────────────────────────
    print("=" * 60)
    print("SEWER OPTIMIZER v2.0")
    print("=" * 60)

    cfg = load_config(args.config)
    print(f"  Config       : {args.config or '(padrões internos)'}")
    print(f"  Moeda        : {cfg.currency} ({cfg.cost_year})")
    print(f"  GA           : pop={cfg.population_size}, gen={cfg.n_generations}, "
          f"Pc={cfg.crossover_rate}, Pm={cfg.mutation_rate}, runs={cfg.n_runs}")

    # ── Carregar rede ─────────────────────────────────────────────────────────
    if not args.network.exists():
        print(f"\nERRO: arquivo de rede não encontrado: {args.network}")
        sys.exit(1)

    print(f"\nCarregando rede: {args.network}")
    edges = load_network_csv(args.network)
    print(f"  {len(edges)} trechos carregados.")

    issues = validate_network(edges)
    if issues:
        print("\nAVISOSS/ERROS na rede:")
        for iss in issues:
            print(f"  [!] {iss}")
        if any("ciclo" in i.lower() or "duplicado" in i.lower() for i in issues):
            print("ABORTAR: erros críticos na topologia.")
            sys.exit(1)

    # ── Topologia e cache ─────────────────────────────────────────────────────
    in_edges, out_edge, topo_order, outlet_node = build_topology(edges)
    print(f"  Nó de saída (outlet): {outlet_node}")
    print(f"  Construindo cache hidráulico "
          f"({len(edges)} trechos × {len(cfg.discrete_diameters_mm)} diâmetros "
          f"× {len(cfg.discrete_slopes)} declividades)...")
    hyd_cache = build_hydraulic_cache(edges, cfg.discrete_diameters_mm, cfg.discrete_slopes)
    print(f"  {len(hyd_cache)} combinações pré-calculadas.")

    # ── Execução do AG ────────────────────────────────────────────────────────
    print(f"\nIniciando {cfg.n_runs} execução(ões) do Algoritmo Genético...")
    stats = run_multiple(edges, cfg, in_edges, out_edge, topo_order, outlet_node, hyd_cache,
                         verbose=not args.quiet)

    # ── Relatório console ─────────────────────────────────────────────────────
    print_run_summary(stats, cfg)

    # ── Exportar resultados ───────────────────────────────────────────────────
    args.output.mkdir(parents=True, exist_ok=True)

    csv_path = export_csv(stats, cfg, args.output)
    print(f"\nCSV exportado   : {csv_path}")

    if not args.no_excel:
        xl_path = export_excel(stats, cfg, args.output)
        if xl_path:
            print(f"Excel exportado : {xl_path}")

    print("\nConcluído.")


if __name__ == "__main__":
    main()
